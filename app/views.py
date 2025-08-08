from django.shortcuts import render

from django.http import HttpResponse
import json
import openpyxl

from .models import Producto, Cliente, Proveedor

# Create your views here.
def Home(request):
    return HttpResponse(json.dumps({'exito':1}), content_type="application/json")

def RegistrarProveedores(request):
    # Specify the path to your XLSX file
    file_path = './filesExcel/Lista_proveedores.xlsx'

    # Load the workbook
    workbook = openpyxl.load_workbook(file_path)

    # Get the active sheet (or a specific sheet by name: workbook['SheetName'])
    sheet = workbook.active

    # Iterate through rows and print cell values
    json_proveedores = []
    cnt = 1
    for row in sheet.iter_rows(min_row=3,min_col=2):
        oProveedor = {
            'cnt': cnt,
            'nombre': row[0].value,
            'direccion': row[1].value,
            'documento': row[2].value,
        }
        cnt += 1
        # Guardar en la base de datos
        if oProveedor['nombre']:
            json_proveedores.append(oProveedor)
            oProveedor = Proveedor(
                nombre = oProveedor['nombre'],
                direccion = oProveedor['direccion'],
                documento = oProveedor['documento'],
            )
            #oProveedor.save()

    return HttpResponse(json.dumps(json_proveedores), content_type="application/json")

def RegistrarClientes(request):
    # Specify the path to your XLSX file
    file_path = './filesExcel/Lista_clientes.xlsx'

    # Load the workbook
    workbook = openpyxl.load_workbook(file_path)

    # Get the active sheet (or a specific sheet by name: workbook['SheetName'])
    sheet = workbook.active

    # Iterate through rows and print cell values
    json_clientes = []
    cnt = 1
    for row in sheet.iter_rows(min_row=3,min_col=2):
        oCliente = {
            'cnt': cnt,
            'nombre': row[0].value,
            'direccion': row[1].value,
            'nro_documento': row[2].value,
            'telefono': row[3].value,
        }
        cnt += 1
        json_clientes.append(oCliente)
        # Guardar en la base de datos
        oCliente = Cliente(
            nombre = oCliente['nombre'],
            direccion = oCliente['direccion'],
            numerodocumento = oCliente['nro_documento'],
            telefono = oCliente['telefono'],
        )
        #oCliente.save()

    return HttpResponse(json.dumps(json_clientes), content_type="application/json")


def RegistrarProductos(request):
    # Specify the path to your XLSX file
    file_path = './filesExcel/Lista_productos.xlsx'

    # Load the workbook
    workbook = openpyxl.load_workbook(file_path)

    # Get the active sheet (or a specific sheet by name: workbook['SheetName'])
    sheet = workbook.active

    # Iterate through rows and print cell values
    json_productos = []
    cnt = 1
    #Nombre Producto	Código	Cant. Mínima	Precio Compra	Precio Venta S/.	Precio Venta $.
    for row in sheet.iter_rows(min_row=3,max_row=1989,min_col=2):
        codigo = row[1].value
        if isinstance(codigo, (int, float)):
            #print("La celda contiene un número:", codigo)
            #print("Valor entero:", valor_entero)
            oProducto = {
                'cnt': cnt,
                'nombre': str(row[0].value),
                'codigo': f"{codigo:.0f}",
                'cant_minima': int(row[2].value),
                'precio_compra': float(row[3].value),
                'precio_venta_soles': float(row[4].value),
                'precio_venta_dolares': float(row[5].value),
            }
        else:
            print("La celda no contiene un número.", codigo)
            oProducto = {
                'cnt': cnt,
                'nombre': str(row[0].value),
                'codigo': codigo,
                'cant_minima': int(row[2].value),
                'precio_compra': float(row[3].value),
                'precio_venta_soles': float(row[4].value),
                'precio_venta_dolares': float(row[5].value),
            }
        
        cnt += 1
        json_productos.append(oProducto)
        # Guardar en la base de datos
        oProducto = Producto(
            nombre = oProducto['nombre'],
            codigo = oProducto['codigo'],
            min_cantidad = oProducto['cant_minima'],
            precio_compra_dolares = oProducto['precio_compra'],
            precio_venta_soles = oProducto['precio_venta_soles'],
            precio_venta_dolares = oProducto['precio_venta_dolares'],
        )
        #oProducto.save()
    return HttpResponse(json.dumps(json_productos), content_type="application/json")
